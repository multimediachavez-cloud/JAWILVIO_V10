import sqlite3
import pandas as pd
import os, re, calendar
import openpyxl
from datetime import datetime, date

from .utils.security import hash_password, is_password_hashed

IMPORT_VERSION = 'v5_socios_dni_foto'

SCHEMA = """
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE NOT NULL,
    password TEXT NOT NULL,
    role TEXT NOT NULL,
    estado TEXT NOT NULL DEFAULT 'Activo',
    creado_en TEXT DEFAULT CURRENT_TIMESTAMP,
    ultimo_acceso TEXT,
    two_factor_enabled INTEGER NOT NULL DEFAULT 0,
    two_factor_secret TEXT
);
CREATE TABLE IF NOT EXISTS socios (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    numero INTEGER UNIQUE,
    nombre TEXT,
    dni TEXT,
    foto TEXT,
    meses INTEGER,
    plazo_balance INTEGER,
    fecha_prestamo TEXT,
    saldo REAL,
    mes_2026 TEXT,
    reunion TEXT,
    permisos TEXT
);
CREATE TABLE IF NOT EXISTS cuotas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    socio_numero INTEGER,
    plazo INTEGER,
    fecha TEXT,
    prestamo REAL,
    interes REAL,
    abono_capital REAL,
    cuota REAL,
    saldo REAL,
    hoja_origen TEXT
);
CREATE TABLE IF NOT EXISTS meta (
    clave TEXT PRIMARY KEY,
    valor TEXT
);
CREATE TABLE IF NOT EXISTS asistencia (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    socio_numero INTEGER,
    socio_nombre TEXT,
    fecha TEXT,
    estado TEXT,
    observacion TEXT
);
CREATE TABLE IF NOT EXISTS multas_asistencia (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    periodo TEXT NOT NULL,
    socio_numero INTEGER NOT NULL,
    socio_nombre TEXT NOT NULL,
    inasistencias INTEGER NOT NULL DEFAULT 0,
    tardanzas INTEGER NOT NULL DEFAULT 0,
    monto_multa_inasistencia REAL NOT NULL DEFAULT 0,
    monto_multa_tardanza REAL NOT NULL DEFAULT 0,
    total_multa REAL NOT NULL DEFAULT 0,
    estado_cobro TEXT NOT NULL DEFAULT 'Pendiente',
    fecha_cobro TEXT,
    observacion TEXT,
    edicion_manual INTEGER NOT NULL DEFAULT 0,
    oculto_manual INTEGER NOT NULL DEFAULT 0,
    creado_en TEXT DEFAULT CURRENT_TIMESTAMP,
    calculado_en TEXT DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(periodo, socio_numero)
);
CREATE TABLE IF NOT EXISTS configuracion (
    clave TEXT PRIMARY KEY,
    valor TEXT
);
CREATE TABLE IF NOT EXISTS auditoria (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    fecha TEXT DEFAULT CURRENT_TIMESTAMP,
    usuario TEXT,
    accion TEXT,
    detalle TEXT,
    categoria TEXT DEFAULT 'usuario',
    modulo TEXT,
    entidad TEXT,
    entidad_id TEXT,
    periodo TEXT,
    nivel TEXT DEFAULT 'INFO',
    antes_json TEXT,
    despues_json TEXT,
    metadata_json TEXT
);
CREATE TABLE IF NOT EXISTS saldo_historico_mensual (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    periodo TEXT NOT NULL,
    socio_numero INTEGER,
    socio_nombre TEXT NOT NULL,
    saldo REAL NOT NULL,
    fuente TEXT,
    UNIQUE(periodo, socio_nombre)
);
CREATE TABLE IF NOT EXISTS periodos (
    periodo TEXT PRIMARY KEY,
    estado TEXT NOT NULL DEFAULT 'Abierto',
    origen TEXT NOT NULL DEFAULT 'app',
    total_socios INTEGER NOT NULL DEFAULT 0,
    total_prestamos REAL NOT NULL DEFAULT 0,
    total_aportes REAL NOT NULL DEFAULT 0,
    total_recaudado REAL NOT NULL DEFAULT 0,
    total_intereses REAL NOT NULL DEFAULT 0,
    total_capital REAL NOT NULL DEFAULT 0,
    total_colocado REAL NOT NULL DEFAULT 0,
    saldo_por_colocar REAL NOT NULL DEFAULT 0,
    aporte_mensual_base REAL NOT NULL DEFAULT 0,
    fecha_apertura TEXT DEFAULT CURRENT_TIMESTAMP,
    fecha_calculo TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS obligaciones_mensuales (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    periodo TEXT NOT NULL,
    socio_numero INTEGER NOT NULL,
    socio_nombre TEXT NOT NULL,
    cuotas INTEGER,
    fecha_prestamo TEXT,
    cuota_plazo INTEGER,
    cuota_fecha TEXT,
    cuota_prestamo REAL NOT NULL DEFAULT 0,
    cuota_interes REAL NOT NULL DEFAULT 0,
    cuota_capital REAL NOT NULL DEFAULT 0,
    aporte_mensual REAL NOT NULL DEFAULT 0,
    total_mes REAL NOT NULL DEFAULT 0,
    saldo_actual REAL NOT NULL DEFAULT 0,
    fuente_saldo TEXT,
    UNIQUE(periodo, socio_numero)
);
CREATE TABLE IF NOT EXISTS obligaciones_mensuales_override (
    periodo TEXT NOT NULL,
    socio_numero INTEGER NOT NULL,
    cuota_prestamo REAL,
    fecha_prestamo TEXT,
    nota TEXT,
    PRIMARY KEY (periodo, socio_numero)
);
CREATE TABLE IF NOT EXISTS aportaciones_mensuales (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    periodo TEXT NOT NULL,
    socio_numero INTEGER NOT NULL,
    socio_nombre TEXT NOT NULL,
    cuotas INTEGER,
    fecha_prestamo TEXT,
    cuota_plazo INTEGER,
    cuota_fecha TEXT,
    cuota_prestamo REAL NOT NULL DEFAULT 0,
    cuota_interes REAL NOT NULL DEFAULT 0,
    cuota_capital REAL NOT NULL DEFAULT 0,
    aporte_mensual REAL NOT NULL DEFAULT 0,
    total_mes REAL NOT NULL DEFAULT 0,
    saldo_actual REAL NOT NULL DEFAULT 0,
    fuente_saldo TEXT,
    UNIQUE(periodo, socio_numero)
);
CREATE TABLE IF NOT EXISTS cierre_mensual (
    periodo TEXT PRIMARY KEY,
    total_socios INTEGER NOT NULL DEFAULT 0,
    total_prestamos REAL NOT NULL DEFAULT 0,
    total_aportes REAL NOT NULL DEFAULT 0,
    total_recaudado REAL NOT NULL DEFAULT 0,
    total_intereses REAL NOT NULL DEFAULT 0,
    total_capital REAL NOT NULL DEFAULT 0,
    aporte_mensual_base REAL NOT NULL DEFAULT 0,
    fecha_calculo TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS reuniones_mensuales (
    periodo TEXT PRIMARY KEY,
    socio_numero INTEGER,
    socio_nombre TEXT,
    estado TEXT NOT NULL DEFAULT 'Pendiente',
    fecha_programada TEXT,
    fecha_realizada TEXT,
    tipo_via TEXT,
    direccion_reunion TEXT,
    observacion TEXT,
    actualizado_por TEXT,
    creado_en TEXT DEFAULT CURRENT_TIMESTAMP,
    actualizado_en TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS permisos_mensuales (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    periodo TEXT NOT NULL,
    socio_numero INTEGER NOT NULL,
    socio_nombre TEXT NOT NULL,
    fecha_permiso TEXT,
    motivo TEXT,
    documento TEXT,
    observacion TEXT,
    registrado_por TEXT,
    creado_en TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS prestamos_nuevos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    periodo TEXT NOT NULL,
    socio_numero INTEGER NOT NULL,
    socio_nombre TEXT NOT NULL,
    monto REAL NOT NULL,
    cuotas INTEGER NOT NULL,
    tasa_mensual REAL NOT NULL DEFAULT 0.01,
    fecha_desembolso TEXT NOT NULL,
    cuota_inicial REAL NOT NULL DEFAULT 0,
    total_interes REAL NOT NULL DEFAULT 0,
    total_pagable REAL NOT NULL DEFAULT 0,
    estado TEXT NOT NULL DEFAULT 'Reservado',
    creado_en TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS prestamos_nuevos_cronograma (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    prestamo_nuevo_id INTEGER NOT NULL,
    plazo INTEGER NOT NULL,
    fecha TEXT,
    prestamo REAL,
    interes REAL,
    abono_capital REAL,
    cuota REAL,
    saldo REAL
);
CREATE TABLE IF NOT EXISTS prioridad_colocacion_manual (
    periodo TEXT NOT NULL,
    socio_numero INTEGER NOT NULL,
    prioridad INTEGER NOT NULL,
    PRIMARY KEY (periodo, socio_numero)
);
CREATE TABLE IF NOT EXISTS historial_prestamos_socios (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    socio_numero INTEGER NOT NULL,
    socio_nombre TEXT NOT NULL,
    prestamo_nuevo_id INTEGER,
    periodo TEXT,
    accion TEXT NOT NULL,
    estado_resultante TEXT,
    monto REAL,
    cuotas INTEGER,
    fecha_desembolso TEXT,
    cuota_inicial REAL,
    total_interes REAL,
    total_pagable REAL,
    saldo_anterior REAL,
    meses_anteriores INTEGER,
    fecha_prestamo_anterior TEXT,
    detalle TEXT,
    creado_por TEXT,
    creado_en TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS historial_prestamos_socios_cuotas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    historial_id INTEGER NOT NULL,
    plazo INTEGER,
    fecha TEXT,
    prestamo REAL,
    interes REAL,
    abono_capital REAL,
    cuota REAL,
    saldo REAL,
    hoja_origen TEXT
);
CREATE TABLE IF NOT EXISTS prestamos_excel_historial (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    socio_numero INTEGER NOT NULL,
    socio_nombre TEXT,
    bloque_orden INTEGER NOT NULL,
    titulo TEXT,
    titulo_manual TEXT,
    saldo_base_ref REAL,
    prestamo_adicional_ref REAL,
    oculto_manual INTEGER NOT NULL DEFAULT 0,
    fecha_inicio_manual TEXT,
    fecha_fin_manual TEXT,
    fecha_inicio TEXT,
    fecha_fin TEXT,
    monto_inicial REAL,
    plazo_total INTEGER,
    interes_total REAL,
    capital_total REAL,
    cuota_total REAL,
    saldo_inicial REAL,
    saldo_final REAL,
    es_activo INTEGER NOT NULL DEFAULT 0,
    hoja_origen TEXT,
    UNIQUE(socio_numero, bloque_orden)
);
CREATE TABLE IF NOT EXISTS prestamos_excel_historial_cuotas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    prestamo_excel_id INTEGER NOT NULL,
    plazo INTEGER,
    fecha TEXT,
    prestamo REAL,
    interes REAL,
    abono_capital REAL,
    cuota REAL,
    saldo REAL,
    hoja_origen TEXT
);
"""

MONTHS = {
    'ENERO': '01',
    'FEBRERO': '02',
    'MARZO': '03',
    'ABRIL': '04',
    'MAYO': '05',
    'JUNIO': '06',
    'JULIO': '07',
    'AGOSTO': '08',
    'SETIEMBRE': '09',
    'SEPTIEMBRE': '09',
    'OCTUBRE': '10',
    'NOVIEMBRE': '11',
    'DICIEMBRE': '12',
}

MONTH_NAMES = {
    1: 'ENERO',
    2: 'FEBRERO',
    3: 'MARZO',
    4: 'ABRIL',
    5: 'MAYO',
    6: 'JUNIO',
    7: 'JULIO',
    8: 'AGOSTO',
    9: 'SETIEMBRE',
    10: 'OCTUBRE',
    11: 'NOVIEMBRE',
    12: 'DICIEMBRE',
}


def connect_db(app):
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    return conn


def init_db(app):
    with connect_db(app) as conn:
        conn.executescript(SCHEMA)
        socios_columns = {row['name'] for row in conn.execute("PRAGMA table_info(socios)").fetchall()}
        if 'dni' not in socios_columns:
            conn.execute('ALTER TABLE socios ADD COLUMN dni TEXT')
        if 'foto' not in socios_columns:
            conn.execute('ALTER TABLE socios ADD COLUMN foto TEXT')
        if 'plazo_balance' not in socios_columns:
            conn.execute('ALTER TABLE socios ADD COLUMN plazo_balance INTEGER')
        reuniones_columns = {row['name'] for row in conn.execute("PRAGMA table_info(reuniones_mensuales)").fetchall()}
        if 'tipo_via' not in reuniones_columns:
            conn.execute('ALTER TABLE reuniones_mensuales ADD COLUMN tipo_via TEXT')
        if 'direccion_reunion' not in reuniones_columns:
            conn.execute('ALTER TABLE reuniones_mensuales ADD COLUMN direccion_reunion TEXT')
        if 'actualizado_por' not in reuniones_columns:
            conn.execute('ALTER TABLE reuniones_mensuales ADD COLUMN actualizado_por TEXT')
        permisos_columns = {row['name'] for row in conn.execute("PRAGMA table_info(permisos_mensuales)").fetchall()}
        if 'motivo' not in permisos_columns:
            conn.execute('ALTER TABLE permisos_mensuales ADD COLUMN motivo TEXT')
        if 'documento' not in permisos_columns:
            conn.execute('ALTER TABLE permisos_mensuales ADD COLUMN documento TEXT')
        multas_columns = {row['name'] for row in conn.execute("PRAGMA table_info(multas_asistencia)").fetchall()}
        if 'edicion_manual' not in multas_columns:
            conn.execute("ALTER TABLE multas_asistencia ADD COLUMN edicion_manual INTEGER NOT NULL DEFAULT 0")
        if 'oculto_manual' not in multas_columns:
            conn.execute("ALTER TABLE multas_asistencia ADD COLUMN oculto_manual INTEGER NOT NULL DEFAULT 0")
        auditoria_columns = {row['name'] for row in conn.execute("PRAGMA table_info(auditoria)").fetchall()}
        if 'categoria' not in auditoria_columns:
            conn.execute("ALTER TABLE auditoria ADD COLUMN categoria TEXT DEFAULT 'usuario'")
        if 'modulo' not in auditoria_columns:
            conn.execute("ALTER TABLE auditoria ADD COLUMN modulo TEXT")
        if 'entidad' not in auditoria_columns:
            conn.execute("ALTER TABLE auditoria ADD COLUMN entidad TEXT")
        if 'entidad_id' not in auditoria_columns:
            conn.execute("ALTER TABLE auditoria ADD COLUMN entidad_id TEXT")
        if 'periodo' not in auditoria_columns:
            conn.execute("ALTER TABLE auditoria ADD COLUMN periodo TEXT")
        if 'nivel' not in auditoria_columns:
            conn.execute("ALTER TABLE auditoria ADD COLUMN nivel TEXT DEFAULT 'INFO'")
        if 'antes_json' not in auditoria_columns:
            conn.execute("ALTER TABLE auditoria ADD COLUMN antes_json TEXT")
        if 'despues_json' not in auditoria_columns:
            conn.execute("ALTER TABLE auditoria ADD COLUMN despues_json TEXT")
        if 'metadata_json' not in auditoria_columns:
            conn.execute("ALTER TABLE auditoria ADD COLUMN metadata_json TEXT")
        users_columns = {row['name'] for row in conn.execute("PRAGMA table_info(users)").fetchall()}
        if 'estado' not in users_columns:
            conn.execute("ALTER TABLE users ADD COLUMN estado TEXT NOT NULL DEFAULT 'Activo'")
        if 'creado_en' not in users_columns:
            conn.execute("ALTER TABLE users ADD COLUMN creado_en TEXT")
        if 'ultimo_acceso' not in users_columns:
            conn.execute("ALTER TABLE users ADD COLUMN ultimo_acceso TEXT")
        if 'two_factor_enabled' not in users_columns:
            conn.execute("ALTER TABLE users ADD COLUMN two_factor_enabled INTEGER NOT NULL DEFAULT 0")
        if 'two_factor_secret' not in users_columns:
            conn.execute("ALTER TABLE users ADD COLUMN two_factor_secret TEXT")
        historial_excel_columns = {row['name'] for row in conn.execute("PRAGMA table_info(prestamos_excel_historial)").fetchall()}
        if 'titulo_manual' not in historial_excel_columns:
            conn.execute('ALTER TABLE prestamos_excel_historial ADD COLUMN titulo_manual TEXT')
        if 'saldo_base_ref' not in historial_excel_columns:
            conn.execute('ALTER TABLE prestamos_excel_historial ADD COLUMN saldo_base_ref REAL')
        if 'prestamo_adicional_ref' not in historial_excel_columns:
            conn.execute('ALTER TABLE prestamos_excel_historial ADD COLUMN prestamo_adicional_ref REAL')
        if 'oculto_manual' not in historial_excel_columns:
            conn.execute('ALTER TABLE prestamos_excel_historial ADD COLUMN oculto_manual INTEGER NOT NULL DEFAULT 0')
        if 'fecha_inicio_manual' not in historial_excel_columns:
            conn.execute('ALTER TABLE prestamos_excel_historial ADD COLUMN fecha_inicio_manual TEXT')
        if 'fecha_fin_manual' not in historial_excel_columns:
            conn.execute('ALTER TABLE prestamos_excel_historial ADD COLUMN fecha_fin_manual TEXT')
        users = [
            ('admin', hash_password('admin123'), 'Administrador'),
            ('tesorero', hash_password('tesorero123'), 'Tesorero'),
            ('secretario', hash_password('secretario123'), 'Secretario'),
            ('consulta', hash_password('consulta123'), 'Consulta'),
        ]
        for user in users:
            conn.execute('INSERT OR IGNORE INTO users(username,password,role) VALUES(?,?,?)', user)
        conn.execute("UPDATE users SET estado='Activo' WHERE estado IS NULL OR TRIM(estado)=''")
        conn.execute("UPDATE users SET creado_en=CURRENT_TIMESTAMP WHERE creado_en IS NULL OR TRIM(creado_en)=''")
        legacy_users = conn.execute("SELECT id, password FROM users").fetchall()
        for legacy_user in legacy_users:
            if not is_password_hashed(legacy_user['password']):
                conn.execute(
                    "UPDATE users SET password=? WHERE id=?",
                    (hash_password(legacy_user['password']), legacy_user['id']),
                )

        config = {
            'nombre_asociacion': 'Asociación JAWILVIO',
            'ubicacion': 'Celendin, Cajamarca',
            'logo_institucional': '',
            'aporte_mensual': '150',
            'saldo_actual_total_oficial': '887838.9',
            'acciones_por_socio_oficial': '29594.6',
            'resumen_financiero_periodo_oficial': '2026-03',
            'total_prestamo_acumulado_oficial': '1444072.7',
            'tasa_prestamo_mensual': '0.01',
            'min_cuotas_prestamo': '12',
            'max_cuotas_prestamo': '84',
            'multa_inasistencia': '10',
            'multa_tardanza': '5',
        }
        for key, value in config.items():
            conn.execute('INSERT OR IGNORE INTO configuracion(clave, valor) VALUES(?,?)', (key, value))
        overrides = [
            ('2026-03', 13, 496.9, None, 'Ajuste histórico oficial de marzo 2026.'),
        ]
        for row in overrides:
            conn.execute(
                '''
                INSERT OR IGNORE INTO obligaciones_mensuales_override(periodo, socio_numero, cuota_prestamo, fecha_prestamo, nota)
                VALUES(?,?,?,?,?)
                ''',
                row,
            )
        conn.commit()


def _to_date(value):
    if pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        if value in ('', '-', 'nan', 'NaT'):
            return None
        if ' - ' in value:
            value = value.split(' - ')[0].strip()
        manual_match = re.fullmatch(r'(\d{1,2})/(\d{1,2})/(\d{4})', value)
        if manual_match:
            day = int(manual_match.group(1))
            month = int(manual_match.group(2))
            year = int(manual_match.group(3))
            if 1 <= month <= 12:
                last_day = calendar.monthrange(year, month)[1]
                day = min(day, last_day)
                return f'{year:04d}-{month:02d}-{day:02d}'
    dt = pd.to_datetime(value, dayfirst=True, errors='coerce')
    if pd.isna(dt):
        return None
    return dt.strftime('%Y-%m-%d')


def _to_float(value):
    if pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.replace('S/', '').replace(',', '').strip()
        if value in ('', '-', 'nan'):
            return None
    try:
        return float(value)
    except Exception:
        return None


def _parse_meses(text):
    if pd.isna(text):
        return None
    match = re.search(r'(\d+)', str(text))
    return int(match.group(1)) if match else None


def _clean_text(value):
    if pd.isna(value):
        return None
    text = str(value).strip()
    return None if text in ('', '-', 'nan', 'NaT') else text


def _parse_numero(value):
    if pd.isna(value):
        return None
    text = str(value).strip()
    match = re.search(r'(\d+)', text)
    if match:
        return int(match.group(1))
    try:
        return int(float(value))
    except Exception:
        return None


def _parse_plazo(value):
    if pd.isna(value):
        return None
    if isinstance(value, str):
        text = value.strip()
        if not re.fullmatch(r'\d+', text):
            return None
        plazo = int(text)
    else:
        try:
            plazo = int(float(value))
        except Exception:
            return None
    return plazo if 0 <= plazo <= 240 else None


def _excel_signature(excel_path):
    stat = os.stat(excel_path)
    return f"{os.path.basename(excel_path)}|{int(stat.st_mtime)}|{stat.st_size}"


def _format_sync_timestamp(value):
    return datetime.fromtimestamp(value).strftime('%Y-%m-%d %H:%M:%S')


def _extract_balance_period(title):
    if not title:
        return None
    text = str(title).upper()
    text = text.replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
    match = re.search(r'MES DE\s+([A-Z]+)\s+DE\s+(\d{4})', text)
    if not match:
        return None
    month = MONTHS.get(match.group(1))
    if not month:
        return None
    return f"{match.group(2)}-{month}"


def _clear_imported_data(conn):
    conn.execute('DELETE FROM cuotas')
    conn.execute('DELETE FROM socios')
    conn.execute('DELETE FROM saldo_historico_mensual')
    conn.execute('DELETE FROM prestamos_excel_historial')
    conn.execute('DELETE FROM prestamos_excel_historial_cuotas')
    conn.execute("DELETE FROM meta WHERE clave IN ('excel_importado', 'excel_signature', 'import_status', 'saldo_historico_2026_03_status', 'import_version', 'excel_sync_at')")


def _snapshot_manual_socios(conn):
    return {
        row['numero']: {
            'dni': row['dni'],
            'foto': row['foto'],
        }
        for row in conn.execute("SELECT numero, dni, foto FROM socios").fetchall()
    }


def _snapshot_manual_prestamos_excel(conn):
    return {
        (row['socio_numero'], row['bloque_orden']): {
            'titulo_manual': row['titulo_manual'],
            'saldo_base_ref': row['saldo_base_ref'],
            'prestamo_adicional_ref': row['prestamo_adicional_ref'],
            'oculto_manual': row['oculto_manual'],
            'fecha_inicio_manual': row['fecha_inicio_manual'],
            'fecha_fin_manual': row['fecha_fin_manual'],
        }
        for row in conn.execute(
            """
            SELECT socio_numero, bloque_orden, titulo_manual, saldo_base_ref, prestamo_adicional_ref,
                   oculto_manual, fecha_inicio_manual, fecha_fin_manual
            FROM prestamos_excel_historial
            """
        ).fetchall()
    }


def _import_socios(conn, excel_path, preserved_socios=None):
    socios_df = pd.read_excel(excel_path, sheet_name='SOCIOS', header=None)
    total = 0
    for _, row in socios_df.iterrows():
        numero = _parse_numero(row.get(0))
        nombre = _clean_text(row.get(1))
        if numero is None or not nombre or str(nombre).upper().startswith('APELLIDOS'):
            continue
        preserved = (preserved_socios or {}).get(numero, {})
        conn.execute(
            '''INSERT OR REPLACE INTO socios(numero,nombre,dni,foto,meses,plazo_balance,fecha_prestamo,saldo,mes_2026,reunion,permisos)
               VALUES(?,?,?,?,?,?,?,?,?,?,?)''',
            (
                numero,
                nombre,
                preserved.get('dni'),
                preserved.get('foto'),
                _parse_meses(row.get(2)),
                None,
                _to_date(row.get(3)),
                _to_float(row.get(4)),
                _clean_text(row.get(5)),
                _to_date(row.get(6)) or _clean_text(row.get(6)),
                _to_date(row.get(7)) or _clean_text(row.get(7)),
            )
        )
        total += 1
    return total


def _select_best_block(blocks, socio_meta):
    if not blocks:
        return []
    if len(blocks) == 1:
        return blocks[0]

    objetivo_fecha = socio_meta['fecha_prestamo'] if socio_meta else None
    objetivo_plazo = (socio_meta['plazo_balance'] or socio_meta['meses']) if socio_meta else None

    def score(block):
        inicio = block[0]
        max_plazo = max([row['plazo'] for row in block], default=0)
        score_value = 0
        if objetivo_fecha and inicio['fecha'] == objetivo_fecha:
            score_value += 1000
        if objetivo_plazo and max_plazo == objetivo_plazo:
            score_value += 200
        if objetivo_plazo and max_plazo <= objetivo_plazo:
            score_value += 50
        score_value += max_plazo
        score_value += int(inicio['fecha'].replace('-', '')) if inicio['fecha'] else 0
        return score_value

    return max(blocks, key=score)


def _normalize_block_title(value):
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r'\s+', ' ', text)
    return text


def _normalized_upper_text(value):
    text = _normalize_block_title(value)
    if not text:
        return ''
    return (
        text.upper()
        .replace('Á', 'A')
        .replace('É', 'E')
        .replace('Í', 'I')
        .replace('Ó', 'O')
        .replace('Ú', 'U')
    )


def _title_from_date(iso_date, fallback=None):
    if not iso_date:
        return fallback
    match = re.fullmatch(r'(\d{4})-(\d{2})-(\d{2})', iso_date)
    if not match:
        return fallback
    year = int(match.group(1))
    month = int(match.group(2))
    return f"{MONTH_NAMES.get(month, str(month))} {year} - PRÉSTAMO"


def _add_months_iso(iso_date, months_to_add):
    if not iso_date:
        return None
    base = datetime.strptime(str(iso_date), '%Y-%m-%d').date()
    month_index = (base.month - 1) + months_to_add
    year = base.year + month_index // 12
    month = (month_index % 12) + 1
    day = min(base.day, calendar.monthrange(year, month)[1])
    return date(year, month, day).isoformat()


def _import_cuotas(conn, excel_path, preserved_historial=None):
    xls = pd.ExcelFile(excel_path)
    hojas = [sheet for sheet in xls.sheet_names if str(sheet).isdigit()]
    total = 0
    for sheet in hojas:
        df = pd.read_excel(excel_path, sheet_name=sheet, header=None)
        socio_meta = conn.execute(
            'SELECT nombre, fecha_prestamo, plazo_balance, meses FROM socios WHERE numero=?',
            (int(sheet),)
        ).fetchone()
        rows = []
        current_title = None
        for _, row in df.iterrows():
            title_cell = _clean_text(row.get(0))
            normalized_title = _normalized_upper_text(title_cell)
            if title_cell and 'PRESTAMO' in normalized_title and 'PLAZO' not in normalized_title and 'SOCIO:' not in normalized_title:
                current_title = _normalize_block_title(title_cell)
            plazo = _parse_plazo(row.get(0))
            fecha = _to_date(row.get(1))
            if plazo is None or fecha is None:
                continue
            rows.append(
                {
                    'plazo': plazo,
                    'fecha': fecha,
                    'prestamo': _to_float(row.get(2)),
                    'interes': _to_float(row.get(3)),
                    'abono_capital': _to_float(row.get(4)),
                    'cuota': _to_float(row.get(5)),
                    'saldo': _to_float(row.get(6)),
                    'titulo': current_title,
                }
            )

        blocks = []
        current_block = []
        seen_positive = False
        for item in rows:
            if item['plazo'] == 0:
                if current_block:
                    blocks.append(current_block)
                current_block = [item]
                seen_positive = False
                continue
            if not current_block:
                current_block = [item]
            elif item['plazo'] == 1 and seen_positive:
                blocks.append(current_block)
                current_block = [item]
            else:
                current_block.append(item)
            if item['plazo'] > 0:
                seen_positive = True
        if current_block:
            blocks.append(current_block)

        selected_block = _select_best_block(blocks, socio_meta)
        selected_signature = tuple((item['plazo'], item['fecha'], item['saldo']) for item in selected_block)
        selected_fecha_inicio_manual = None
        selected_fecha_inicio_resuelta = None

        for block_index, block in enumerate(blocks, start=1):
            block_signature = tuple((item['plazo'], item['fecha'], item['saldo']) for item in block)
            is_active = 1 if block_signature == selected_signature else 0
            first_item = block[0] if block else {}
            positive_rows = [item for item in block if (item.get('plazo') or 0) > 0]
            existing_historial = (preserved_historial or {}).get((int(sheet), block_index), {})
            block_title = first_item.get('titulo') or _title_from_date(first_item.get('fecha'), f'Prestamo {block_index}')
            fecha_inicio_manual = existing_historial.get('fecha_inicio_manual')
            fecha_fin_manual = existing_historial.get('fecha_fin_manual')
            fecha_inicio_resuelta = fecha_inicio_manual or first_item.get('fecha')
            fecha_fin_resuelta = fecha_fin_manual or (
                _add_months_iso(fecha_inicio_resuelta, max([item['plazo'] for item in positive_rows], default=0))
                if fecha_inicio_resuelta and positive_rows
                else (block[-1].get('fecha') if block else None)
            )
            historial_cursor = conn.execute(
                """
                INSERT OR REPLACE INTO prestamos_excel_historial(
                    socio_numero, socio_nombre, bloque_orden, titulo, titulo_manual, saldo_base_ref, prestamo_adicional_ref,
                    oculto_manual, fecha_inicio_manual, fecha_fin_manual, fecha_inicio, fecha_fin,
                    monto_inicial, plazo_total, interes_total, capital_total, cuota_total,
                    saldo_inicial, saldo_final, es_activo, hoja_origen
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    int(sheet),
                    socio_meta['nombre'] if socio_meta else None,
                    block_index,
                    block_title,
                    existing_historial.get('titulo_manual'),
                    existing_historial.get('saldo_base_ref'),
                    existing_historial.get('prestamo_adicional_ref'),
                    existing_historial.get('oculto_manual', 0) or 0,
                    fecha_inicio_manual,
                    fecha_fin_manual,
                    fecha_inicio_resuelta,
                    fecha_fin_resuelta,
                    first_item.get('prestamo'),
                    max([item['plazo'] for item in positive_rows], default=0),
                    round(sum([(item.get('interes') or 0) for item in positive_rows]), 1),
                    round(sum([(item.get('abono_capital') or 0) for item in positive_rows]), 1),
                    round(sum([(item.get('cuota') or 0) for item in positive_rows]), 1),
                    first_item.get('saldo'),
                    block[-1].get('saldo') if block else None,
                    is_active,
                    str(sheet),
                ),
            )
            prestamo_excel_id = historial_cursor.lastrowid
            if not prestamo_excel_id:
                existing = conn.execute(
                    'SELECT id FROM prestamos_excel_historial WHERE socio_numero=? AND bloque_orden=?',
                    (int(sheet), block_index),
                ).fetchone()
                prestamo_excel_id = existing['id'] if existing else None
            if is_active:
                selected_fecha_inicio_manual = fecha_inicio_manual
                selected_fecha_inicio_resuelta = fecha_inicio_resuelta
            for item in block:
                fecha_item = _add_months_iso(fecha_inicio_resuelta, item['plazo']) if fecha_inicio_manual else item['fecha']
                conn.execute(
                    """
                    INSERT INTO prestamos_excel_historial_cuotas(
                        prestamo_excel_id, plazo, fecha, prestamo, interes, abono_capital, cuota, saldo, hoja_origen
                    ) VALUES(?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        prestamo_excel_id,
                        item['plazo'],
                        fecha_item,
                        item['prestamo'],
                        item['interes'],
                        item['abono_capital'],
                        item['cuota'],
                        item['saldo'],
                        str(sheet),
                    ),
                )

        for item in selected_block:
            fecha_item = _add_months_iso(selected_fecha_inicio_resuelta, item['plazo']) if selected_fecha_inicio_manual else item['fecha']
            conn.execute(
                '''INSERT INTO cuotas(socio_numero, plazo, fecha, prestamo, interes, abono_capital, cuota, saldo, hoja_origen)
                   VALUES(?,?,?,?,?,?,?,?,?)''',
                (
                    int(sheet),
                    item['plazo'],
                    fecha_item,
                    item['prestamo'],
                    item['interes'],
                    item['abono_capital'],
                    item['cuota'],
                    item['saldo'],
                    str(sheet),
                )
            )
            total += 1
        if selected_block and selected_fecha_inicio_resuelta:
            conn.execute(
                'UPDATE socios SET fecha_prestamo=? WHERE numero=?',
                (selected_fecha_inicio_resuelta, int(sheet)),
            )
    return total, hojas


def _import_balance_history(conn, excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if 'BALANCE' not in wb.sheetnames:
        return None, 0, 0
    ws = wb['BALANCE']
    periodo = _extract_balance_period(ws['A1'].value) or '2026-03'
    total = 0
    vinculados = 0
    for row_idx in range(4, ws.max_row + 1):
        numero = _parse_numero(ws.cell(row_idx, 1).value)
        nombre = _clean_text(ws.cell(row_idx, 2).value)
        plazo_balance = _parse_meses(ws.cell(row_idx, 3).value)
        fecha_balance = _to_date(ws.cell(row_idx, 4).value)
        saldo = _to_float(ws.cell(row_idx, 9).value)
        if numero is None or not nombre or saldo is None:
            continue
        exists = conn.execute('SELECT 1 FROM socios WHERE numero=?', (numero,)).fetchone()
        if exists:
            vinculados += 1
            conn.execute(
                '''
                UPDATE socios
                SET plazo_balance=COALESCE(?, plazo_balance),
                    fecha_prestamo=COALESCE(fecha_prestamo, ?)
                WHERE numero=?
                ''',
                (plazo_balance, fecha_balance, numero)
            )
        conn.execute(
            '''INSERT OR REPLACE INTO saldo_historico_mensual(periodo, socio_numero, socio_nombre, saldo, fuente)
               VALUES(?,?,?,?,?)''',
            (periodo, numero, nombre, saldo, 'balance_excel')
        )
        total += 1
    return periodo, total, vinculados


def import_excel_if_needed(app):
    excel_path = app.config['EXCEL_PATH']
    with connect_db(app) as conn:
        if not os.path.exists(excel_path):
            conn.execute("INSERT OR REPLACE INTO meta(clave, valor) VALUES('import_status', ?)", ('Archivo Excel no encontrado',))
            conn.commit()
            return

        signature = _excel_signature(excel_path)
        current_signature = conn.execute("SELECT valor FROM meta WHERE clave='excel_signature'").fetchone()
        already_imported = conn.execute('SELECT COUNT(*) FROM socios').fetchone()[0] > 0
        balance_plazos_cargados = conn.execute(
            'SELECT COUNT(*) FROM socios WHERE plazo_balance IS NOT NULL'
        ).fetchone()[0]
        missing_prestamo_dates = conn.execute(
            "SELECT COUNT(*) FROM socios WHERE fecha_prestamo IS NULL OR TRIM(fecha_prestamo) = '' OR fecha_prestamo = '-'"
        ).fetchone()[0]
        current_import_version = conn.execute(
            "SELECT valor FROM meta WHERE clave='import_version'"
        ).fetchone()
        current_sync_at = conn.execute(
            "SELECT valor FROM meta WHERE clave='excel_sync_at'"
        ).fetchone()
        if (
            already_imported
            and current_signature
            and current_signature['valor'] == signature
            and balance_plazos_cargados > 0
            and missing_prestamo_dates == 0
            and current_import_version
            and current_import_version['valor'] == IMPORT_VERSION
        ):
            if not current_sync_at:
                conn.execute(
                    "INSERT OR REPLACE INTO meta(clave, valor) VALUES('excel_sync_at', ?)",
                    (_format_sync_timestamp(os.path.getmtime(excel_path)),)
                )
                conn.commit()
            return

        preserved_socios = _snapshot_manual_socios(conn)
        preserved_historial = _snapshot_manual_prestamos_excel(conn)
        _clear_imported_data(conn)
        _import_socios(conn, excel_path, preserved_socios)
        cuotas_insertadas, hojas = _import_cuotas(conn, excel_path, preserved_historial)
        periodo_balance, balance_insertados, vinculados = _import_balance_history(conn, excel_path)
        socios_insertados = conn.execute('SELECT COUNT(*) FROM socios').fetchone()[0]

        conn.execute("INSERT OR REPLACE INTO meta(clave, valor) VALUES('excel_importado', ?)", (os.path.basename(excel_path),))
        conn.execute("INSERT OR REPLACE INTO meta(clave, valor) VALUES('excel_signature', ?)", (signature,))
        conn.execute("INSERT OR REPLACE INTO meta(clave, valor) VALUES('import_version', ?)", (IMPORT_VERSION,))
        conn.execute("INSERT OR REPLACE INTO meta(clave, valor) VALUES('excel_sync_at', ?)", (datetime.now().strftime('%Y-%m-%d %H:%M:%S'),))
        conn.execute(
            "INSERT OR REPLACE INTO meta(clave, valor) VALUES('import_status', ?)",
            (f'Socios: {socios_insertados} | Cuotas: {cuotas_insertadas} | Hojas: {len(hojas)} | Balance: {balance_insertados}',)
        )
        if periodo_balance:
            conn.execute(
                "INSERT OR REPLACE INTO meta(clave, valor) VALUES('saldo_historico_2026_03_status', ?)",
                (f'Registros: {balance_insertados} | Vinculados: {vinculados}',)
            )
        conn.execute(
            """
            INSERT INTO auditoria(
                usuario, accion, detalle, categoria, modulo, entidad, periodo, nivel, metadata_json
            ) VALUES(?,?,?,?,?,?,?,?,?)
            """,
            (
                'sistema',
                'Importacion Excel',
                f'Socios {socios_insertados}, cuotas {cuotas_insertadas}, hojas {len(hojas)}, balance {balance_insertados}',
                'sistema',
                'importacion_excel',
                'excel',
                periodo_balance,
                'INFO',
                (
                    '{'
                    f'"socios": {int(socios_insertados)}, '
                    f'"cuotas": {int(cuotas_insertadas)}, '
                    f'"hojas": {int(len(hojas))}, '
                    f'"balance": {int(balance_insertados)}'
                    '}'
                ),
            )
        )
        conn.commit()


def seed_saldos_historicos(app):
    return
